"""
Tool Calling 테스트 공통 유틸리티
- run_test : spy 패턴으로 tool 선택 검증 + 타이밍 측정
- run_group: 케이스 목록 실행 후 결과 출력
- print_summary: 그룹별/전체 요약 출력
- save_xlsx : 결과를 xlsx 파일로 저장
"""
import time
import os
from datetime import datetime
from unittest.mock import patch
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import app.services.llm as llm_module

load_dotenv()

_token = os.getenv("TEST_JWT_TOKEN", "")
USER_CONTEXT = {"user_id": 202, "account_id": 1, "token": _token}


def run_test(case: dict) -> dict:
    """
    단일 케이스 실행.

    타이밍:
        time_to_tool_ms  : chat() 시작 → spy_dispatch 최초 호출 (LLM 도구 선택 결정까지)
        time_to_answer_ms: spy_dispatch 최초 호출 → chat() 반환 (도구 실행 + 답변 생성)
        total_ms         : chat() 전체 소요 시간
    """
    captured = []
    original_dispatch = llm_module._dispatch
    t_tool_selected: float | None = None

    def spy_dispatch(fn_name, fn_args, user_context):
        nonlocal t_tool_selected
        if t_tool_selected is None:
            t_tool_selected = time.perf_counter()
        captured.append({"tool": fn_name, "args": fn_args})
        return original_dispatch(fn_name, fn_args, user_context)

    t_start = time.perf_counter()
    with patch.object(llm_module, "_dispatch", side_effect=spy_dispatch):
        try:
            reply = llm_module.chat(case["message"], USER_CONTEXT)
        except Exception as e:
            t_end = time.perf_counter()
            return {
                "status":            "ERROR",
                "error":             str(e),
                "captured":          captured,
                "reply":             "",
                "actual_tool":       None,
                "actual_type":       None,
                "actual_args":       {},
                "time_to_tool_ms":   None,
                "time_to_answer_ms": None,
                "total_ms":          round((t_end - t_start) * 1000),
            }
    t_end = time.perf_counter()

    time_to_tool_ms   = round((t_tool_selected - t_start) * 1000) if t_tool_selected else None
    time_to_answer_ms = round((t_end - t_tool_selected) * 1000)   if t_tool_selected else None
    total_ms          = round((t_end - t_start) * 1000)

    actual_tool = captured[0]["tool"] if captured else None
    actual_args = captured[0]["args"] if captured else {}
    actual_type = actual_args.get("type") if captured else None

    # 판정
    if case["expected_tool"] is None:
        ok = actual_tool is None
    else:
        tool_ok = actual_tool == case["expected_tool"]
        type_ok = actual_type == case["expected_type"]
        args_ok = True
        if case["expected_args"]:
            args_ok = all(actual_args.get(k) == v for k, v in case["expected_args"].items())
        ok = tool_ok and type_ok and args_ok

    return {
        "status":            "PASS" if ok else "FAIL",
        "actual_tool":       actual_tool,
        "actual_type":       actual_type,
        "actual_args":       actual_args,
        "reply":             reply,
        "captured":          captured,
        "time_to_tool_ms":   time_to_tool_ms,
        "time_to_answer_ms": time_to_answer_ms,
        "total_ms":          total_ms,
    }


def run_group(test_cases: list[dict]) -> list[dict]:
    """케이스 목록을 순서대로 실행하고 결과를 출력 + 반환."""
    results = []
    current_group = None

    for case in test_cases:
        if case["group"] != current_group:
            current_group = case["group"]
            print(f"\n{'━'*60}")
            print(f"  {current_group}")
            print(f"{'━'*60}")

        print(f"\n  Q: {case['message']}")
        print(f"     ({case['desc']})")

        res = run_test(case)

        exp_str = (f"{case['expected_tool']} / type={case['expected_type']}"
                   if case["expected_tool"] else "tool 미호출")
        act_str = (f"{res['actual_tool']} / type={res['actual_type']}"
                   if res["actual_tool"] else "tool 미호출")

        print(f"     기대: {exp_str}")
        print(f"     실제: {act_str}")

        if case["expected_args"] and res["actual_args"]:
            for k, v in case["expected_args"].items():
                actual_v = res["actual_args"].get(k)
                mark = "✓" if actual_v == v else "✗"
                print(f"     args.{k}: 기대={v}, 실제={actual_v} {mark}")

        # 타이밍
        if res["time_to_tool_ms"] is not None:
            print(f"     시간  도구선택: {res['time_to_tool_ms']}ms"
                  f"  |  답변생성: {res['time_to_answer_ms']}ms"
                  f"  |  합계: {res['total_ms']}ms")
        else:
            print(f"     시간  합계: {res['total_ms']}ms  (tool 미호출)")

        print(f"     [{res['status']}]", end="")
        if res["status"] == "ERROR":
            print(f" {res['error']}", end="")
        print()

        if res.get("reply"):
            snippet = res["reply"][:300]
            suffix  = "..." if len(res["reply"]) > 300 else ""
            print(f"     답변: {snippet}{suffix}")

        results.append({**case, **res})

    return results


def print_summary(results: list[dict]) -> None:
    """그룹별 + 전체 PASS/FAIL/ERROR 및 평균 타이밍 출력."""
    print(f"\n{'='*60}")
    print("  결과 요약")
    print(f"{'='*60}")

    groups: dict[str, list] = {}
    for r in results:
        groups.setdefault(r["group"], []).append(r)

    total_pass = total_fail = total_error = 0
    for group, items in groups.items():
        p = sum(1 for i in items if i["status"] == "PASS")
        f = sum(1 for i in items if i["status"] == "FAIL")
        e = sum(1 for i in items if i["status"] == "ERROR")
        total_pass += p
        total_fail += f
        total_error += e

        tool_times   = [i["time_to_tool_ms"]   for i in items if i["time_to_tool_ms"]   is not None]
        answer_times = [i["time_to_answer_ms"]  for i in items if i["time_to_answer_ms"] is not None]
        avg_tool   = f"{round(sum(tool_times)   / len(tool_times))}ms"   if tool_times   else "-"
        avg_answer = f"{round(sum(answer_times) / len(answer_times))}ms" if answer_times else "-"

        print(f"  {group:<32} PASS {p}/{len(items)}", end="")
        if f:
            print(f"  FAIL {f}", end="")
        if e:
            print(f"  ERROR {e}", end="")
        print(f"  |  도구선택 평균 {avg_tool} / 답변생성 평균 {avg_answer}")

    print(f"{'─'*60}")
    total = len(results)
    print(f"  전체: PASS {total_pass}/{total}  FAIL {total_fail}  ERROR {total_error}")

    path = save_xlsx(results)
    print(f"\n  xlsx 저장 완료: {path}")


def save_xlsx(results: list[dict], out_dir: str = "test_results") -> str:
    """테스트 결과를 xlsx로 저장하고 파일 경로를 반환합니다."""
    os.makedirs(out_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"test_result_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── 스타일 정의 ──────────────────────────────────────────────
    fill_pass    = PatternFill("solid", fgColor="C6EFCE")   # 연초록
    fill_fail    = PatternFill("solid", fgColor="FFC7CE")   # 연빨강
    fill_error   = PatternFill("solid", fgColor="FFEB9C")   # 연노랑
    fill_header  = PatternFill("solid", fgColor="4472C4")   # 파랑
    fill_group   = PatternFill("solid", fgColor="D9E1F2")   # 연파랑
    font_header  = Font(bold=True, color="FFFFFF")
    font_group   = Font(bold=True)
    font_fail    = Font(bold=True, color="9C0006")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = [
        "그룹", "설명", "질문 메시지",
        "기대 도구", "기대 타입",
        "실제 도구", "실제 타입",
        "결과",
        "도구선택(ms)", "답변생성(ms)", "합계(ms)",
        "LLM 답변",
    ]
    COL_WIDTHS = [28, 22, 30, 16, 18, 16, 18, 8, 13, 13, 10, 60]

    # ── 시트 1: 전체 결과 ────────────────────────────────────────
    ws = wb.active
    ws.title = "전체 결과"

    # 헤더
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = fill_header
        cell.font   = font_header
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # 데이터 행
    groups: dict[str, list] = {}
    for r in results:
        groups.setdefault(r["group"], []).append(r)

    row_idx = 2
    for group, items in groups.items():
        for item in items:
            status = item["status"]
            values = [
                item.get("group", ""),
                item.get("desc", ""),
                item.get("message", ""),
                item.get("expected_tool") or "미호출",
                item.get("expected_type") or "-",
                item.get("actual_tool")   or "미호출",
                item.get("actual_type")   or "-",
                status,
                item.get("time_to_tool_ms")   or "-",
                item.get("time_to_answer_ms") or "-",
                item.get("total_ms")          or "-",
                item.get("reply", "")[:500],  # 500자 제한
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = align_left if col in (1, 2, 3, 12) else align_center
                cell.border = border
                if status == "PASS":
                    cell.fill = fill_pass
                elif status == "FAIL":
                    cell.fill = fill_fail
                    if col in (6, 7, 8):
                        cell.font = font_fail
                elif status == "ERROR":
                    cell.fill = fill_error
            ws.row_dimensions[row_idx].height = 40
            row_idx += 1

    # ── 시트 2: 그룹별 요약 ─────────────────────────────────────
    ws2 = wb.create_sheet("그룹별 요약")
    sum_headers = ["그룹", "전체", "PASS", "FAIL", "ERROR", "도구선택 평균(ms)", "답변생성 평균(ms)"]
    sum_widths  = [32, 8, 8, 8, 8, 18, 18]

    for col, (h, w) in enumerate(zip(sum_headers, sum_widths), start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill   = fill_header
        cell.font   = font_header
        cell.alignment = align_center
        cell.border = border
        ws2.column_dimensions[get_column_letter(col)].width = w

    row_idx2 = 2
    total_p = total_f = total_e = 0
    for group, items in groups.items():
        p = sum(1 for i in items if i["status"] == "PASS")
        f = sum(1 for i in items if i["status"] == "FAIL")
        e = sum(1 for i in items if i["status"] == "ERROR")
        total_p += p; total_f += f; total_e += e

        tool_times   = [i["time_to_tool_ms"]   for i in items if i["time_to_tool_ms"]   is not None]
        answer_times = [i["time_to_answer_ms"]  for i in items if i["time_to_answer_ms"] is not None]
        avg_tool   = round(sum(tool_times)   / len(tool_times))   if tool_times   else "-"
        avg_answer = round(sum(answer_times) / len(answer_times)) if answer_times else "-"

        row_vals = [group, len(items), p, f, e, avg_tool, avg_answer]
        for col, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx2, column=col, value=val)
            cell.alignment = align_center if col != 1 else align_left
            cell.border = border
            cell.fill = fill_group
            if col == 1:
                cell.font = font_group
        ws2.row_dimensions[row_idx2].height = 18
        row_idx2 += 1

    # 합계 행
    total_row = ["합계", len(results), total_p, total_f, total_e, "-", "-"]
    for col, val in enumerate(total_row, start=1):
        cell = ws2.cell(row=row_idx2, column=col, value=val)
        cell.font   = Font(bold=True)
        cell.alignment = align_center if col != 1 else align_left
        cell.border = border
        cell.fill   = PatternFill("solid", fgColor="BDD7EE")

    wb.save(path)
    return path
