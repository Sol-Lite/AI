import os
import re
import sys
import time
import html
import certifi
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import (
    MONGO_URI, DB_NAME, COLLECTION_NAME,
    HEADERS, NEWS_API_URL, ARTICLE_URL,
    MEDIA_KEYWORDS, TARGET_PER_STOCK, SCHEDULE_INTERVAL,
)
from summarizer import summarize_articles
from bs4 import BeautifulSoup
from datetime import datetime
from pymongo import MongoClient
import openpyxl

# ── MongoDB 연결 ──────────────────────────────────────────────
client = MongoClient(MONGO_URI, tlsCAFile=certifi.where())
db = client[DB_NAME]
collection = db[COLLECTION_NAME]

# 인덱스 (최초 1회, 스키마 기준)
collection.create_index("news_id", unique=True, sparse=True)
collection.create_index([("published_at", -1)])
collection.create_index([("related_stocks.stock_code", 1)])


# ── 본문 전처리 ───────────────────────────────────────────────
_KNOWN_MEDIA_SN = (
    r'뉴시스|뉴스1|연합뉴스|파이낸셜뉴스|헤럴드경제|이데일리|머니투데이'
    r'|한국경제|조선일보|중앙일보|동아일보|한겨레|매일경제|이코노미스트'
    r'|데일리안|더팩트|서울신문|서울경제|아시아경제|아이뉴스|뉴스핌'
)

_LEADING_BYLINE_RE_SN = re.compile(
    r'^(?:'
    r'[\[\(][^\]\)]{1,60}(?:기자|특파원|앵커|논설위원|선임기자|보험전문기자)[\]\)]\s*'
    r'|'
    r'[\[\(][^\]\)]{1,50}[\]\)]\s*[가-힣A-Za-z·\s]{0,20}'
    r'(?:기자|특파원|앵커|보험전문기자|논설위원|선임기자)\s*[=:]?\s*[가-힣]{0,15}\s*'
    r'|'
    r'\[(?:' + _KNOWN_MEDIA_SN + r')\]\s*'
    r')',
    re.DOTALL,
)


def _strip_leading_bylines(text: str) -> str:
    """본문 앞 기자/언론사 바이라인을 건너뛰고 실제 내용 시작점부터 반환"""
    while True:
        m = _LEADING_BYLINE_RE_SN.match(text)
        if not m:
            break
        text = text[m.end():].lstrip()
    return text


def clean_body(text):
    # 0. 본문 앞 바이라인 건너뛰기
    text = _strip_leading_bylines(text)
    # 1. 기자 바이라인 제거 (문두 — 위에서 못 잡은 나머지 fallback)
    text = re.sub(r'^\[[^\]]{1,40}기자\]\s*', '', text)
    text = re.sub(r'^\[[가-힣a-zA-Z=\s]{2,20}\]\s*[가-힣\s]{2,15}기자\s*=\s*', '', text)
    text = re.sub(r'^\([가-힣a-zA-Z0-9=\s]{2,20}\)\s*[가-힣\s]{2,15}기자\s*=\s*', '', text)
    text = re.sub(r'^\[[가-힣a-zA-Z\s=|ㅣ]{2,20}\]\s*', '', text)
    # 2. 방송 아티팩트 제거
    text = re.sub(r'재생\[[^\]]{1,20}\]', '', text)
    text = re.sub(r'◀\s*(앵커|리포트)\s*▶\s*', '', text)
    text = re.sub(r'<\s*(앵커|기자|리포트)\s*>', '', text)
    text = re.sub(r'\[\s*(앵커|리포트|기자)\s*\]', '', text)
    # 3. ━ 이후 한글 나오기 전까지 제거
    text = re.sub(r'━[^가-힣]*', '', text)
    # 4. 인라인 사진/제공/표 표기 제거
    text = re.sub(r'\(사진[^)]{0,30}\)', '', text)
    text = re.sub(r'\(제공[^)]{0,20}\)', '', text)
    text = re.sub(r'\(표[^)]{0,20}\)', '', text)
    text = re.sub(r'<사진>', '', text)
    # 5. 기자 이메일 포함 대괄호 제거 (예: [김혜인 디지털팀 기자 haileykim@gmail.com])
    text = re.sub(r'\[[^\]]*기자[^\]]*@[^\]]*\]', '', text)
    # 6. 기자명 + 이메일 패턴 제거 (예: 박두호 기자 walnut@etnews.com)
    text = re.sub(r'[가-힣]{2,6}\s*기자\s+[a-zA-Z0-9._+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', text)
    # 7. 단독 이메일 주소 제거 (한글 앞에 붙은 것 제외)
    text = re.sub(r'[a-zA-Z0-9_+-][a-zA-Z0-9._+-]*@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', text)
    # 8. URL 제거
    text = re.sub(r'https?://\S+', '', text)
    # 9. 매체 제보 안내 블록 제거 (▶카카오톡/▶이메일/▶뉴스 홈페이지 등)
    text = re.sub(r'▶.{0,30}(카카오톡|제보|홈페이지|이메일).{0,100}', '', text)
    # 10. 장식용 특수문자 제거
    text = re.sub(r'[■◆★☆◇○□]+\s*', '', text)
    # 11. 캡션 텍스트 [출처] 패턴 (앞 텍스트까지 통째로 제거)
    _SRC = (
        r'제공|SNS|AP|EPA|AFP|로이터|Reuters|게티|Getty'
        r'|뉴시스|뉴스1|연합뉴스|연합|뉴스핌|이데일리|머니투데이'
        r'|한국경제|조선일보|중앙일보|동아일보|한겨레|매일경제|헤럴드경제'
    )
    _pat = r'[^\[]{2,80}\s*\[[^\]]*(?:' + _SRC + r')[^\]]*\]\s*'
    text = re.sub(_pat, ' ', text)
    # 12. 캡션 텍스트 [언론사=기자] / [언론사 | 기자] 패턴 (앞 짧은 텍스트 포함 제거)
    text = re.sub(r'[가-힣\w·,·\s]{2,60}\[[^\]]{2,40}기자\]\s*', ' ', text)
    # 13. 남은 [기자] 바이라인 제거
    text = re.sub(r'\[[^\]]{1,50}기자\]\s*', '', text)
    text = re.sub(r'\[[^\]\|]{1,30}\|[^\]]{1,30}기자\]\s*', '', text)
    # 14. 사진 | 브랜드명 캡션 크레딧 제거
    text = re.sub(r'\s*사진\s*[|=]\s*[^\s][^.。\n]{0,30}', '', text)
    # 15. ▲ 로 시작하는 프로모션 문구 제거 (AI 프리즘 등)
    text = re.sub(r'▲\s*.{0,200}', '', text)
    # 16. 연속 공백 정리
    return re.sub(r'\s+', ' ', text).strip()


# ── 종목 로드 ─────────────────────────────────────────────────
def load_kospi200(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    code_idx = headers.index("종목코드")
    name_idx = headers.index("종목명")
    return [
        {"ticker": str(row[code_idx]).zfill(6), "name": row[name_idx], "market": "KOSPI"}
        for row in rows[1:] if row[code_idx]
    ]


def load_nasdaq100(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    sym_idx = headers.index("Symbol")
    name_idx = headers.index("Name")
    return [
        {"ticker": f"{row[sym_idx]}.O", "name": row[name_idx], "market": "NASDAQ"}
        for row in rows[1:] if row[sym_idx]
    ]


# ── 기사 전문 파싱 ────────────────────────────────────────────
def fetch_article_body(office_id, article_id):
    url = ARTICLE_URL.format(officeId=office_id, articleId=article_id)
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # 대표 이미지
        og_image = soup.select_one('meta[property="og:image"]')
        thumbnail_url = og_image["content"] if og_image else ""

        content = soup.select_one("div#dic_area") or soup.select_one("div.newsct_article")
        if not content:
            return "", [], thumbnail_url

        def _extract_lines(tag):
            """<br> 태그를 줄바꿈으로 처리 후 비어있지 않은 줄 반환"""
            for br in tag.select("br"):
                br.replace_with("\n")
            return [s.strip() for s in tag.get_text(separator="\n").split("\n") if s.strip()]

        subtitles = []

        # 방식 1: strong.media_end_summary (요약형 소제목)
        summary_tag = content.select_one("strong.media_end_summary")
        if summary_tag:
            subtitles += _extract_lines(summary_tag)
            summary_tag.decompose()

        # 방식 2: border-left div/strong — 인라인 스타일 소제목
        for tag in content.select("div[style*='border-left'], strong[style*='border-left']"):
            for line in _extract_lines(tag):
                if len(line) > 5 and line not in subtitles:
                    subtitles.append(line)
            tag.decompose()

        # 방식 3: b 태그 소제목 (border-left 내부 b는 이미 위에서 처리)
        for b_tag in content.select("b"):
            if b_tag.find_parent(style=lambda s: s and "border-left" in s):
                continue
            text = b_tag.get_text(strip=True)
            if text and len(text) > 5 and text not in subtitles:
                subtitles.append(text)
            b_tag.decompose()

        for tag in content.select("span._PHOTO_VIEWER, em.img_desc"):
            tag.decompose()

        return clean_body(content.get_text(strip=True)), subtitles, thumbnail_url
    except Exception as e:
        print(f"    본문 파싱 실패 ({office_id}/{article_id}): {e}")
    return "", [], ""


# ── 중복 체크 ─────────────────────────────────────────────────
def deduplicate(articles):
    """
    1단계: 현재 배치 내 news_id 기준 중복 제거
    2단계: DB 내 기존 기사와 news_id 기준 중복 제거
    """
    # 1단계 - 배치 내 중복
    seen = set()
    unique = []
    for a in articles:
        if a["news_id"] not in seen:
            seen.add(a["news_id"])
            unique.append(a)

    # 2단계 - DB 중복
    news_ids = [a["news_id"] for a in unique]
    existing = set(
        doc["news_id"] for doc in collection.find({"news_id": {"$in": news_ids}}, {"news_id": 1})
    )
    return [a for a in unique if a["news_id"] not in existing]


# ── 종목별 크롤링 ─────────────────────────────────────────────
def crawl_stock_news(stock, target=TARGET_PER_STOCK):
    """
    유효 기사(본문 50자 이상)가 target개 될 때까지 페이지를 넘겨 가며 크롤링.
    배치 내 중복은 link 기준으로 실시간 제거.
    """
    ticker = stock["ticker"]
    valid = []
    seen_links = set()
    page = 1
    fetch_size = target  # 한 번에 요청할 기사 수

    while len(valid) < target:
        url = NEWS_API_URL.format(ticker=ticker, page_size=fetch_size, page=page)
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"    [{ticker}] API 요청 실패 (page {page}): {e}")
            break

        if not data or "items" not in data[0] or not data[0]["items"]:
            break  # 더 이상 기사 없음

        items = data[0]["items"]

        for item in items:
            if len(valid) >= target:
                break

            title = html.unescape(item.get("titleFull") or item.get("title", ""))

            # 사진/영상 기사 필터
            if any(kw in title for kw in MEDIA_KEYWORDS):
                continue

            office_id = item.get("officeId", "")
            article_id = item.get("articleId", "")
            link = ARTICLE_URL.format(officeId=office_id, articleId=article_id)

            news_id = f"{office_id}_{article_id}"

            # 배치 내 중복 스킵
            if news_id in seen_links:
                continue
            seen_links.add(news_id)

            # 전문 파싱
            body, subtitles, thumbnail_url = fetch_article_body(office_id, article_id)
            if len(body) < 50:
                # 본문 없음 → 카운트하지 않고 다음 기사로
                time.sleep(0.3)
                continue

            valid.append({
                "news_id": f"{office_id}_{article_id}",
                "title": title,
                "subtitles": subtitles,
                "content": body,
                "thumbnail_url": thumbnail_url,
                "source": item.get("officeName", ""),
                "source_url": link,
                "related_stocks": [
                    {"stock_code": ticker, "stock_name": stock["name"], "market": stock["market"]}
                ],
                "published_at": datetime.strptime(item["datetime"], "%Y%m%d%H%M")
                                if item.get("datetime") else datetime.now(),
                "fetched_at": datetime.now(),
            })
            time.sleep(0.3)

        # 한 페이지에서 가져온 기사가 fetch_size보다 적으면 더 이상 없음
        if len(items) < fetch_size:
            break

        page += 1

    return valid


# ── 1회 사이클 ────────────────────────────────────────────────
def run_job(stocks):
    print(f"\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] 크롤링 사이클 시작 (총 {len(stocks)}종목, 목표 {TARGET_PER_STOCK}건/종목)")
    all_new = []

    for stock in stocks:
        ticker = stock["ticker"]

        # 크롤링 (본문 없는 기사는 카운트 제외, 부족하면 추가 페이지 fetch)
        candidates = crawl_stock_news(stock, target=TARGET_PER_STOCK)

        if not candidates:
            print(f"  [{ticker}] {stock['name']} — 새 기사 없음")
            time.sleep(0.5)
            continue

        # 2. 중복 체크 (배치 내 + DB)
        new_articles = deduplicate(candidates)

        if not new_articles:
            print(f"  [{ticker}] {stock['name']} — 모두 중복")
            time.sleep(0.5)
            continue

        # 3. 요약
        new_articles = summarize_articles(new_articles)

        # 4. MongoDB 저장
        collection.insert_many(new_articles)
        print(f"  [{ticker}] {stock['name']} — {len(new_articles)}건 저장 완료")
        all_new.extend(new_articles)

        time.sleep(0.5)

    print(f"사이클 완료. 총 신규 기사: {len(all_new)}건")


# ── 메인 ─────────────────────────────────────────────────────
if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    kospi_stocks = load_kospi200(os.path.join(base_dir, "코스피200리스트.xlsx"))
    nasdaq_stocks = load_nasdaq100(os.path.join(base_dir, "Nasdaq-100.xlsx"))
    all_stocks = kospi_stocks + nasdaq_stocks

    print(f"KOSPI {len(kospi_stocks)}종목 + NASDAQ {len(nasdaq_stocks)}종목 = 총 {len(all_stocks)}종목")
    print(f"30분 주기 크롤링 시작 (종목당 유효 기사 목표: {TARGET_PER_STOCK}건)")

    while True:
        try:
            run_job(all_stocks)
        except Exception as e:
            print(f"루프 에러: {e}")
        print(f"{SCHEDULE_INTERVAL // 60}분 대기 중...")
        time.sleep(SCHEDULE_INTERVAL)
